import json
from datetime import timedelta, datetime

from django.http import JsonResponse, HttpResponse
from django.utils.timezone import now, make_aware
from django.views.generic import TemplateView, CreateView
from django.core.exceptions import ValidationError
from openpyxl.workbook import Workbook

from robots.forms import DateForm
from robots.models import Robot


class IndexView(TemplateView):
    """
    Отображение главной страницы.
    """
    template_name = "robots/index.html"


class RobotCreateView(CreateView):
    """
    Отображение страницы для создания одного объекта класса Robot через форму.
    """
    model = Robot
    fields = ['model', 'version', 'created']
    template_name = "robots/add_robot_form.html"

    def form_invalid(self, form):
        """Возвращает ошибки валидации в формате JSON."""
        return JsonResponse({"success": False, "errors": form.errors}, status=400)

    def form_valid(self, form):
        """Сохраняет форму и возвращает успешный ответ."""
        robot = form.save(commit=False)
        robot.serial = f"{robot.model}-{robot.version}"
        robot.save()
        return JsonResponse({"success": True, "robot_id": robot.id})


class RobotBulkCreateView(CreateView):
    """
    Отображение страницы для создания одного или нескольких объектов класса Robot через ввод данных в формате JSON.
    """
    model = Robot
    fields = ['model', 'version', 'created']
    template_name = "robots/add_robots_json_form.html"

    def post(self, request, *args, **kwargs):
        try:
            # Получаем JSON-данные из текстового поля
            data = json.loads(request.POST.get("robots_json", "[]"))

            # Если данные — один объект, оборачиваем в массив
            if isinstance(data, dict):  # Один объект
                data = [data]

            robots_to_create = []
            errors = []

            for item in data:
                try:
                    robot = Robot(
                        model=item.get("model").upper(),
                        version=item.get("version").upper(),
                        created=item.get("created")
                    )
                    robot.serial = f"{robot.model}-{robot.version}"
                    robot.full_clean()  # Валидация данных
                    robots_to_create.append(robot)
                except ValidationError as e:
                    errors.append({"data": item, "errors": e.message_dict})

            # Сохраняем валидные объекты
            if robots_to_create:
                Robot.objects.bulk_create(robots_to_create)

            return JsonResponse({
                "success": len(robots_to_create),
                "errors": errors,
            })
        except json.JSONDecodeError as err:
            return JsonResponse({"error": f"Invalid JSON format: {str(err)}"}, status=400)


class ExcelReportView(TemplateView):
    """
    Отображение страницы для ввода даты и получения отчета в excel-файл.
    """
    form_class = DateForm
    template_name = "robots/excel_report_form.html"

    def get_context_data(self, **kwargs):
        """Переопределение метода для отображения нужной формы."""
        context = super().get_context_data(**kwargs)
        context["form"] = self.form_class()
        return context


def export_to_excel(request):
    """
    Функция для получения данных из формы, фильтрации данных в БД и создания excel-файла.
    """
    # Обработка формы
    form = DateForm(request.GET or None)
    if form.is_valid():
        end_date = form.cleaned_data['date']
    else:
        end_date = now().date()

    # Расчет начала периода
    start_date = end_date - timedelta(days=7)
    start_datetime = make_aware(datetime.combine(start_date, datetime.min.time()))
    end_datetime = make_aware(datetime.combine(end_date, datetime.max.time()))

    # Фильтрация объектов за последнюю неделю
    robots = Robot.objects.filter(created__range=(start_datetime, end_datetime))

    # Группировка по модели и версии
    grouped_data = {}
    for robot in robots:
        if robot.model not in grouped_data:
            grouped_data[robot.model] = {}
        if robot.version not in grouped_data[robot.model]:
            grouped_data[robot.model][robot.version] = 0
        grouped_data[robot.model][robot.version] += 1

    if not grouped_data:
        return JsonResponse({"message": "Нет данных за указанный период"}, status=404)

    # Создание Excel-файла
    wb = Workbook()
    for model, versions in grouped_data.items():
        ws = wb.create_sheet(title=model)
        ws.append(['Модель', 'Версия', 'Количество за неделю'])
        for version, count in versions.items():
            ws.append([model, version, count])

    # Удаляем стандартный пустой лист, если он существует
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Создание HTTP-ответа с Excel-файлом
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=robots_{end_date}.xlsx'
    wb.save(response)
    return response
